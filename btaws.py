#!/usr/bin/env python3

from openpyxl import Workbook
from boto3 import client
import pickle
from os import path
import string
from operator import add
from functools import reduce
from commandr import command, Run

cache_resources = True


class cached_property(object):

    def __init__(self, func):
        self.__doc__ = getattr(func, '__doc__')
        self.func = func
        self.cached_property_name = '__cached__{}__'.format(func.__name__)
        self.file_path = path.join('/tmp', func.__name__)

    def __get__(self, obj, cls):

        if self.cached_property_name in obj.__dict__:
            return obj.__dict__[self.cached_property_name]

        if cache_resources and path.exists(self.file_path):
            return pickle.load(open(self.file_path, 'rb'))

        value = obj.__dict__[self.cached_property_name] = self.func(obj)

        if cache_resources:
            pickle.dump(value, open(self.file_path, 'wb'))

        return value


def normalize_tags(obj):
    obj['TagsByName'] = {t['Key']: t['Value'] for t in obj['Tags']}


class View(object):

    def __init__(self, enviroment, scope):
        self.ec2 = client('ec2', region_name='eu-west-1')
        self.r53 = client('route53', region_name='eu-west-1')
        self.filters = [{'Name': 'tag:Scope', 'Values': [scope]}] if scope else None
        self.scope = scope
        self.environment = enviroment
        self.image_cache = {}

    @cached_property
    def instances(self):
        instances = []
        next_token = None
        while True:
            kwargs = {'NextToken': next_token} if next_token else {}
            if self.filters:
                kwargs['Filters'] = self.filters

            response = self.ec2.describe_instances(**kwargs)
            for reservation in response.get('Reservations', []):
                for instance in reservation.get('Instances', []):
                    instances.append(instance)
                    normalize_tags(instance)

            next_token = response.get('NextToken')

            if not next_token:
                break
        return instances

    @cached_property
    def subnets(self):
        return self.ec2.describe_subnets().get('Subnets')

    @cached_property
    def route53_records(self):
        records = []
        hosted_zones = self.r53.list_hosted_zones().get('HostedZones')
        for hosted_zone in hosted_zones:
            record_sets = self.r53 \
                .list_resource_record_sets(HostedZoneId=hosted_zone.get('Id')) \
                .get('ResourceRecordSets')

            for record in record_sets:
                records.append(record)

        return records

    @cached_property
    def security_groups(self):
        return self.ec2.describe_security_groups().get('SecurityGroups')

    def debug(self):
        from pprint import pprint

        # pprint(self.instances[0])
        # print(self.subnets)
        # print(self.route53_records)
        # print(self.security_groups)

        # for r in self.route53_records:
        #     print(r)

        return self

    def get_dns_for_ip(self, ip):
        matching_record = {}
        for record in self.route53_records:
            for record_value in record.get('ResourceRecords', []):
                if record_value['Value'] == ip:
                    matching_record = record
                    break

        return matching_record

    def get_subnet(self, subnet_id):
        subnets = {s['SubnetId']: s for s in self.subnets}
        subnet = subnets[subnet_id]
        normalize_tags(subnet)
        return subnet

    def describe_images(self, image_ids=[]):
        if not image_ids:
            return {}

        images = self.ec2.describe_images(ImageIds=image_ids)['Images']
        return {i['ImageId']: i['Name'] for i in images}

    def write(self, target_file):
        wb = Workbook()

        self._add_info(wb.active)
        self._add_instances(wb.create_sheet())

        wb.save(target_file)

    def _add_info(self, ws):
        ws.title = 'Info'

    def _add_instances(self, ws):
        ws.title = 'Instances'
        instances = self.instances
        instances = sorted(instances, key=lambda instance: instance['TagsByName'].get('Name', ''))

        image_ids = list(set(i['ImageId'] for i in instances))
        image_names = self.describe_images(image_ids)

        tags = map(lambda instance: list(instance['TagsByName'].keys()), instances)
        tags = set(reduce(add, tags, []))
        tags = filter(lambda tag: not tag.upper().startswith('CEC'), tags)
        tags = sorted(tags)

        columns = [
            ('InstanceId', None),
            ('InstanceType', None),
            ('ImageId', None),
            ('ImageId', lambda id: image_names.get(id, 'Outdated WIN')),
            ('Placement', lambda value: value['AvailabilityZone']),
            ('PublicIpAddress', None),
            ('PrivateIpAddress', None),
            ('PrivateIpAddress', lambda ip: self.get_dns_for_ip(ip).get('Name'), 'DNS Name'),
            # ('LaunchTime', lambda value: 'TIME'),
            ('State', lambda value: value['Name']),
            ('SubnetId', None),
            ('SubnetId', lambda id: self.get_subnet(id)['TagsByName']['Name'], 'Subnet Name'),
            ('SubnetId', lambda id: self.get_subnet(id)['CidrBlock'], 'Subnet Cidr'),
        ]

        # Set headers
        for letter, (prop, _, *args) in zip(string.ascii_uppercase, columns):
            ws['{}{}'.format(letter, 1)] = args[0] if args else prop

        for letter, tag_name in zip(string.ascii_letters[len(columns):], tags):
            ws['{}{}'.format(letter, 1)] = 'Tag: {}'.format(tag_name)

        # Set values
        for row, instance in enumerate(instances):
            row += 2
            for letter, (prop, get, *args) in zip(string.ascii_uppercase, columns):
                get = get or (lambda x: x)
                value = get(instance.get(prop)) or ''
                ws['{}{}'.format(letter, row)] = value

            for letter, tag_name in zip(string.ascii_letters[len(columns):], tags):
                ws['{}{}'.format(letter, row)] = instance['TagsByName'].get(tag_name)


@command('create')
def create_dump(enviroment, scope, output_file, dev_mode=False):
    global cache_resources
    cache_resources = dev_mode

    assert scope in (None, 'LevelB', 'LevelC'), 'Should be one of LevelB or LevelC'

    target_file = path.realpath(output_file)
    View(enviroment, scope).debug().write(target_file)


def run():
    Run()

if __name__ == '__main__':
    Run()
